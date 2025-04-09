import pandas as pd
import numpy as np
from openpyxl import load_workbook
from constants import PRICELIST_PATH, LINK_COLUMN_NAME

def retrieve_xlsx() -> pd.DataFrame:
    
    # Считываем данные через pandas
    df = pd.read_excel(PRICELIST_PATH, engine='openpyxl')
    
    # Загружаем книгу через openpyxl в режиме read_only=False
    wb = load_workbook(PRICELIST_PATH, read_only=False, data_only=False)
    ws = wb.active
    
    # Указываем индекс столбца с гиперссылками (начиная с 0)
    link_column_idx = 1  # Пример для столбца C
    
    hyperlinks = []
    for row in ws.iter_rows(min_row=2, max_row=len(df)+1):  # Обрабатываем только строки с данными
        cell = row[link_column_idx]
        
        # Проверяем наличие стандартной гиперссылки
        if cell.hyperlink:
            hyperlinks.append(cell.hyperlink.target)
        # Если гиперссылка добавлена через формулу HYPERLINK
        elif cell.data_type == 'f' and "HYPERLINK" in cell.value:
            formula = cell.value
            # Извлекаем URL из формулы (пример: =HYPERLINK("http://example.com","Текст"))
            url = formula.split('"')[1] if '"' in formula else None
            hyperlinks.append(url)
        else:
            hyperlinks.append(None)
    
    # Добавляем ссылки в DataFrame
    df[LINK_COLUMN_NAME] = hyperlinks
    
    return process_dataframe(df)

def process_dataframe(df) -> pd.DataFrame:
    # Создаем новый столбец для категорий
    df['Категория'] = np.nan
    
    current_category = None
    
    # Проходим по всем строкам датафрейма
    for index, row in df.iterrows():
        # Проверяем, является ли строка категорией (если третий столбец NaN)
        if pd.isna(row.iloc[1]):
            current_category = row.iloc[0].lower()  # Сохраняем название категории в нижнем регистре
        else:
            df.at[index, 'Категория'] = current_category
    
    # Удаляем строки с категориями (где третий столбец NaN)
    df = df[~pd.isna(df.iloc[:, 1])]

    # BEGIN: Filter columns
    df = df[['арт', 'Наименование', 'Страна', 'Розничная цена за 1 кг/шт', 'Ед. изм.', LINK_COLUMN_NAME, 'Категория']]
    # END: Filter columns
    # Преобразуем все данные в столбцах в нижний регистр, если это строковые значения
    df = df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
    
    # Сбрасываем индексы
    df.reset_index(drop=True, inplace=True)
    
    return df
